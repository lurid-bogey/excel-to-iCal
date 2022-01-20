
import sys
import io
import os
import os.path
import datetime
import argparse
from typing import List

try:
    import openpyxl
except ImportError:
    print(f'*** Error: Failed to import the "openpyxl" package.')
    sys.exit(-1)

from version import VERSION, DESCRIPTION


INPUT_FOLDER = 'in'
OUTPUT_FOLDER = 'out'
DOMAIN = 'lurid_bogey_ical_generator.dummy.local'


def createUniqueIdentifier(eventName: str, eventYear: str) -> str:
    # The UID should be unique, but not random.
    # Subsequent runs should deterministically produce the same UID for the same event.
    # Once again: the UID should be unique, but not random!
    eventName = eventName.replace(' ', '_')
    uid = f'{eventYear}-{eventName}@{DOMAIN}'
    return uid


def createOneEvent(uid: str, date: datetime.datetime, summary: str, descr: str) -> str:
    """Create an iCal event from a simple template.

    :param uid: UID of the event.
    :param date: All our events are one-day events. This is the start date of the event. The time is irrelevant.
                 The end date is calculated from the start date (end = start + 1).
    :param summary: The name of the event.
    :param descr: Optional description.
    :return: One iCal event.
    """
    dateStart = date.date().isoformat().replace('-', '')
    dateEnd = date + datetime.timedelta(days=1)
    dateEnd = dateEnd.date().isoformat().replace('-', '')

    if descr is not None:
        descr = descr.replace('\n', '\\n')

    event = f"""BEGIN:VEVENT
UID:{uid}
TRANSP:TRANSPARENT
X-APPLE-TRAVEL-ADVISORY-BEHAVIOR:AUTOMATIC
DTSTART;VALUE=DATE:{dateStart}
DTEND;VALUE=DATE:{dateEnd}
SUMMARY:{summary}
DESCRIPTION:{descr}
END:VEVENT"""

    if descr is None:
        event = event.replace('DESCRIPTION:None\n', '')

    return event


def enumerateExcelFiles(folder: str) -> List[str]:
    """Return a list of all Excel files in the folder.

    :param folder: Folder
    :return: List of file names
    """
    excelFiles = []
    try:
        entities = os.listdir(folder)
    except FileNotFoundError:
        print(f'*** Error: Folder "{folder}" not found.')
    else:
        for entity in entities:
            if (os.path.isfile(os.path.join(folder, entity)) and
                    entity.endswith('xlsx') and not
                    entity.startswith('~')):
                excelFiles.append(os.path.join(folder, entity))

    return excelFiles


def main(inputFolder=INPUT_FOLDER, outputFolder=OUTPUT_FOLDER):
    # collect the Excel files from the input folder
    excelFiles = enumerateExcelFiles(folder=inputFolder)
    # iterate over all Excel files
    for excelFile in excelFiles:
        print(f'Processing "{excelFile}"...')

        # determine the output file name and make the output folder, if needed
        outFileName = os.path.split(excelFile)[1].replace('xlsx', 'ics')
        if not os.path.isdir(outputFolder):
            os.makedirs(outputFolder)
        outFileName = os.path.join(outputFolder, outFileName)

        # open the file for writing
        with io.open(outFileName, 'w', encoding='utf8') as outFile:

            # the start sequence
            outFile.write('BEGIN:VCALENDAR\nVERSION:2.0\n')

            # open the Excel workbook
            wb = openpyxl.load_workbook(filename=excelFile, data_only=True)
            sheetnames = wb.sheetnames

            # iterate over all sheets and columns
            for sheet in sheetnames:
                ws = wb[sheet]
                for col in ws.columns:
                    # create one iCal event per cell and write it to the output file.
                    # the first cell contains the name of the event.
                    # the second cell contains the description.
                    eventName = col[0].value
                    try:
                        assert isinstance(eventName, str)
                    except AssertionError:
                        print(f'*** Error in cell "{col[0].coordinate}". Aborting...')
                        print(f'*** The cell must contain an event name.')
                        sys.exit(-1)

                    eventDescr = col[1].value
                    for cell in col[2:]:
                        eventDate = cell.value
                        try:
                            assert isinstance(eventDate, datetime.datetime)
                        except AssertionError:
                            print(f'*** Error in cell "{cell.coordinate}". Aborting...')
                            print(f'*** The cell must contain a date.')
                            sys.exit(-1)
                        else:
                            eventYear = str(eventDate.year)
                            uid = createUniqueIdentifier(eventName=eventName, eventYear=eventYear)
                            icalEvent = createOneEvent(uid=uid, date=eventDate, summary=eventName, descr=eventDescr)

                            outFile.write(icalEvent)
                            outFile.write('\n')

            # close the Excel workbook and write the end sequence
            wb.close()
            outFile.write('END:VCALENDAR\n')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(prog='excel-to-iCal.py', description='Generate iCal files')
    parser.add_argument('-i', '--in', dest='input_Folder', default=INPUT_FOLDER, help="Folder with Excel files")
    parser.add_argument('-o', '--out', dest='output_Folder', default=OUTPUT_FOLDER, help="Destination folder for iCal files")
    parser.add_argument('-v', '--version', action="store_true", default=None, help="Shows version and exits")
    args = parser.parse_args()

    if args.version:
        print(DESCRIPTION, VERSION)
        sys.exit(0)
    else:
        main(inputFolder=args.input_Folder, outputFolder=args.output_Folder)
        print('Done')
