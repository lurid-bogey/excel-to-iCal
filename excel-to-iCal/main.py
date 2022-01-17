
import sys
import io
import os
import os.path
import datetime
import openpyxl
from typing import List

INPUT_FOLDER = 'in'
OUTPUT_FOLDER = 'out'
DOMAIN = 'lurid_bogey_ical_generator.dummy.local'


def createUniqueIdentifier(eventName: str, eventYear: str) -> str:
    # FIXME: Pay attention to your input! It could produce non-unique UID :(
    eventName = eventName.replace(' ', '_')
    uid = f'{eventYear}-{eventName}@{DOMAIN}'
    return uid


def createOneEvent(uid: str, date: datetime.datetime, summary: str, descr: str) -> str:
    """Create an iCal event from a simple template.

    :param uid: UID of the event.
    :param date: All our events are one-day events. This is the start date of the event. The time is irrelevant.
                 The end date is calculated from the start date (end = start + 1).
    :param summary: The name of the event.
    :param descr: Optional description.
    :return: One iCal event.
    """
    dateStart = date.date().isoformat().replace('-', '')
    dateEnd = date + datetime.timedelta(days=1)
    dateEnd = dateEnd.date().isoformat().replace('-', '')

    if descr is not None:
        descr = descr.replace('\n', '\\n')

    event = f"""BEGIN:VEVENT
UID:{uid}
TRANSP:TRANSPARENT
X-APPLE-TRAVEL-ADVISORY-BEHAVIOR:AUTOMATIC
DTSTART;VALUE=DATE:{dateStart}
DTEND;VALUE=DATE:{dateEnd}
SUMMARY:{summary}
DESCRIPTION:{descr}
END:VEVENT"""

    if descr is None:
        event = event.replace('DESCRIPTION:None\n', '')

    return event


def enumerateExcelFiles(folder: str) -> List[str]:
    """Return a list of all Excel files in the folder.

    :param folder: Folder
    :return: List of file names
    """
    excelFiles = []
    entities = os.listdir(folder)
    for entity in entities:
        if (os.path.isfile(os.path.join(folder, entity)) and
                entity.endswith('xlsx') and not
                entity.startswith('~')):
            excelFiles.append(os.path.join(folder, entity))

    return excelFiles


if __name__ == '__main__':
    # collect the Excel files from the input folder
    excelFiles = enumerateExcelFiles(folder=INPUT_FOLDER)
    # iterate over all Excel files
    for excelFile in excelFiles:

        # determine the output file name
        outFileName = os.path.split(excelFile)[1].replace('xlsx', 'ics')
        outFileName = os.path.join(OUTPUT_FOLDER, outFileName)
        # open the file for writing
        with io.open(outFileName, 'w', encoding='utf8') as outFile:

            # Write the start sequence
            outFile.write('BEGIN:VCALENDAR\nVERSION:2.0\n')

            # open the Excel workbook
            wb = openpyxl.load_workbook(filename=excelFile, data_only=True)
            sheetnames = wb.sheetnames

            # iterate over all sheets and columns
            for sheet in sheetnames:
                ws = wb[sheet]
                for col in ws.columns:
                    # create one iCal event per cell and write it to the output file
                    # the first cell contains the name of the event
                    # the second cell contains the description
                    eventName = col[0].value
                    eventDescr = col[1].value
                    for cell in col[2:]:
                        eventDate = cell.value
                        try:
                            assert isinstance(eventDate, datetime.datetime)
                        except AssertionError:
                            print(f'Error in {cell.coordinate}. Aborting...')
                            sys.exit(-1)
                        else:
                            eventYear = str(eventDate.year)
                            uid = createUniqueIdentifier(eventName=eventName, eventYear=eventYear)
                            icalEvent = createOneEvent(uid=uid, date=eventDate, summary=eventName, descr=eventDescr)

                            outFile.write(icalEvent)
                            outFile.write('\n')

            # close the Excel workbook and write the end sequence
            wb.close()
            outFile.write('END:VCALENDAR\n')
